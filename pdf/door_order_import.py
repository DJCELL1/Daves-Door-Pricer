import openpyxl


COL_DOORNO = 1
COL_DESC = 2
COL_HAND = 3
COL_UNDERCUT = 4
COL_TRIM_W = 5
COL_TRIM_H = 6
COL_LEAF_W = 7
COL_LEAF_H = 8
COL_92 = 9
COL_112 = 10
COL_136 = 11
COL_SINGLE = 12
COL_DOUBLE = 13
COL_SLIDE = 14


def read_order_form(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    start_row = 14
    results = []

    row = start_row

    while True:
        door_no = ws.cell(row, COL_DOORNO).value

        # End when empty row encountered
        if door_no in (None, ""):
            break

        # Determine jamb size from tick
        jamb = ""
        if ws.cell(row, COL_92).value == "✓":
            jamb = "US14 92x18 Undershot"
        elif ws.cell(row, COL_112).value == "✓":
            jamb = "US13 112x18 Undershot"
        elif ws.cell(row, COL_136).value == "✓":
            jamb = "DG1 136x30 Double Grooved"
        else:
            jamb = ""  # None mapped

        # Determine door type
        if ws.cell(row, COL_SINGLE).value == "✓":
            form = "Single"
        elif ws.cell(row, COL_DOUBLE).value == "✓":
            form = "Double"
        else:
            form = ""

        results.append({
            "Door #": door_no,
            "Room": ws.cell(row, COL_DESC).value,
            "Handing": ws.cell(row, COL_HAND).value,
            "UnderCut": ws.cell(row, COL_UNDERCUT).value,
            "LeafWidth": ws.cell(row, COL_LEAF_W).value,
            "LeafHeight": ws.cell(row, COL_LEAF_H).value,
            "JambType": jamb,
            "Form": form,
        })

        row += 1

    return results
