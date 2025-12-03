import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from io import BytesIO
import datetime

# Column index constants
COL_DOORNO = 1
COL_DESC = 2
COL_HAND = 3
COL_UNDERCUT = 4
COL_TRIM_W = 5
COL_TRIM_H = 6
COL_LEAF_W = 7
COL_LEAF_H = 8
COL_92 = 9
COL_112 = 10
COL_136 = 11
COL_SINGLE = 12
COL_DOUBLE = 13
COL_SLIDE = 14


def safe_write(ws, cell_addr, value):
    """Write to a cell, redirecting to the anchor if the cell is part of a merged range."""
    cell = ws[cell_addr]

    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                anchor = merged_range.coord.split(":")[0]  # top-left cell
                ws[anchor].value = value
                return
    else:
        cell.value = value


def tick(cell):
    cell.value = "✓"
    cell.alignment = Alignment(horizontal="center", vertical="center")


def generate_order_form(template_path, job_details, contractor_details, door_rows):

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ==========================================
    # SAFE HEADER FIELD WRITES
    # ==========================================

    safe_write(ws, "C4", job_details.get("quote", ""))
    safe_write(ws, "C5", job_details.get("project", ""))
    safe_write(ws, "C6", job_details.get("address", ""))

    safe_write(ws, "H4", contractor_details.get("contractor", ""))
    safe_write(ws, "H5", contractor_details.get("contact", ""))
    safe_write(ws, "H6", contractor_details.get("phone", ""))
    safe_write(ws, "H7", contractor_details.get("email", ""))
    safe_write(ws, "H8", contractor_details.get("onsite", ""))

    safe_write(ws, "H9", datetime.date.today().strftime("%d/%m/%Y"))

    # ==========================================
    # WRITE DOOR TABLE
    # ==========================================

    start_row = 14

    for i, d in enumerate(door_rows):
        r = start_row + i

        ws.cell(r, COL_DOORNO).value = d.get("Door #", "")
        ws.cell(r, COL_DESC).value = d.get("Room", "")
        ws.cell(r, COL_HAND).value = d.get("Handing", "")
        ws.cell(r, COL_UNDERCUT).value = d.get("UnderCut", "")

        ws.cell(r, COL_TRIM_W).value = ""   # always blank
        ws.cell(r, COL_TRIM_H).value = ""   # always blank

        ws.cell(r, COL_LEAF_W).value = d.get("LeafWidth")
        ws.cell(r, COL_LEAF_H).value = d.get("LeafHeight")

        # Jamb ticks
        jt = d.get("JambType", "").lower()
        if "92x18" in jt:
            tick(ws.cell(r, COL_92))
        if "112x18" in jt:
            tick(ws.cell(r, COL_112))
        if "136x30" in jt:
            tick(ws.cell(r, COL_136))

        # Single/Double
        form = d.get("Form", "").lower()
        if form == "single":
            tick(ws.cell(r, COL_SINGLE))
        elif form == "double":
            tick(ws.cell(r, COL_DOUBLE))

    # ==========================================
    # RETURN BINARY CONTENTS
    # ==========================================

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
