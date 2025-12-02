import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime


HDL_ORANGE = "FF6600"


def generate_production_template(df_quote, client, project, quote_number):
    """
    Builds an XLSX template duplicated per Qty for site measurements,
    with HDL logo + corporate orange styling.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Production Template"

    # ----------------------------------------------------
    # INSERT HDL LOGO
    # ----------------------------------------------------
    try:
        logo = Image("mnt/data/Logos-01.png")  # adjust if path differs
        logo.width = 200
        logo.height = 60
        ws.add_image(logo, "A1")
    except Exception as e:
        print("Logo load failed:", e)

    # ----------------------------------------------------
    # TITLE BAR
    # ----------------------------------------------------
    ws.merge_cells("A3:F3")
    title_cell = ws["A3"]
    title_cell.value = "HDL PRODUCTION MEASUREMENT TEMPLATE"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=HDL_ORANGE, end_color=HDL_ORANGE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ----------------------------------------------------
    # HEADER DETAILS
    # ----------------------------------------------------
    bold = Font(bold=True)

    headers = [
        ("Client Name:", client),
        ("Project:", project),
        ("Quote Number:", quote_number),
        ("Generated On:", datetime.now().strftime("%d-%m-%Y"))
    ]

    row = 5
    for label, value in headers:
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = bold
        ws[f"B{row}"].value = value
        row += 1

    # ----------------------------------------------------
    # TABLE HEADERS
    # ----------------------------------------------------
    table_headers = [
        "Door Number",
        "SKU",
        "Description",
        "Height",
        "Width",
        "Form",
        "Qty",
        "Undercut (mm)",
        "Finished Floor Height (mm)"
    ]

    start_row = row + 2

    for col, header in enumerate(table_headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HDL_ORANGE, end_color=HDL_ORANGE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ----------------------------------------------------
    # DUPLICATE ROWS PER QTY
    # ----------------------------------------------------
    write_row = start_row + 1

    for _, r in df_quote.iterrows():
        for i in range(r["Qty"]):
            ws.cell(write_row, 1, "")  # Door Number
            ws.cell(write_row, 2, r["SKU"])
            ws.cell(write_row, 3, r["Description"])
            ws.cell(write_row, 4, r["Height"])
            ws.cell(write_row, 5, r["Width"])
            ws.cell(write_row, 6, r["Form"])
            ws.cell(write_row, 7, 1)  # Always 1 per row
            ws.cell(write_row, 8, "")
            ws.cell(write_row, 9, "")
            write_row += 1

    # ----------------------------------------------------
    # AUTO-COLUMN WIDTH
    # ----------------------------------------------------
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    # ----------------------------------------------------
    # EXPORT AS BYTES
    # ----------------------------------------------------
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
